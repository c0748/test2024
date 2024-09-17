import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

# 元データのExcelファイルを読み込む
source_wb = load_workbook("files/colored_sequence_PLOT2.xlsx")
source_ws = source_wb.active

# 新しいExcelファイルを作成
new_wb = Workbook()
new_ws = new_wb.active

# 転記先の表のヘッダーを設定
new_ws.append(["ID", "スタート時間", "終了時間", "サイクルタイム"])

# データを転記
row_id = 1
for i in range(1, source_ws.max_row + 1, 4):
    start_time = source_ws.cell(row=i, column=2).value
    end_time = source_ws.cell(row=i+3, column=2).value
    
    # サイクルタイムを計算
    cycle_time = f"=C{row_id + 1}-B{row_id + 1}"
    
    # 新しいワークシートにデータを追加
    new_ws.append([row_id, start_time, end_time, cycle_time])
    row_id += 1

# 新しいファイルに保存
new_output_path = "files/colored_sequence_PLOT3.xlsx"
new_wb.save(new_output_path)
