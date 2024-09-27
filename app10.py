# PLOT用.xlsxを使って正常サイクルに色を付ける
# 色を付けたファイルをcolored_sequence_PLOT.xlsx保存する
# 正常サイクルだけを抜き出したファイルをcolored_sequence_PLOT2保存する

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl
from datetime import datetime

wb = load_workbook("files/PLOT用.xlsx")
ws = wb.active

# 新しいワークブックを作成
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# シーケンスの色を定義
sequence_color = "FFFF00"  # 黄色

# 必要なシーケンスを定義
required_sequence = ["巻_1", "巻_2", "切_1", "切_2-1"]

# C列の日付ごとに処理するための辞書を作成
date_groups = {}
for row in range(2, ws.max_row + 1):
    date_value = ws.cell(row=row, column=3).value  # C列（3列目）の日付を取得
    if isinstance(date_value, datetime):  # 日付型かどうかを確認
        date_value = date_value.date()  # 日付部分のみを取得（時間情報は無視）
        if date_value not in date_groups:
            date_groups[date_value] = []  # 新しい日付があればキーとして登録
        date_groups[date_value].append(row)  # 日付に対応する行を追加

# 各日付ごとに処理を実行
for date, rows in date_groups.items():
    row_idx = 0
    while row_idx <= len(rows) - 4:  # 最低3行あるか確認
        # 4行のシーケンスを取得
        current_sequence = [ws.cell(row=rows[row_idx + i], column=1).value for i in range(4)]

        # シーケンスが指定されたものと一致するかチェック
        if current_sequence == required_sequence:
            # 一致した場合、その4行に色を適用
            fill = PatternFill(start_color=sequence_color, end_color=sequence_color, fill_type="solid")
            for i in range(4):
                for cell in ws[rows[row_idx + i]]:
                    cell.fill = fill

            # 新しいワークブックに一致した行をコピー
            for i in range(4):
                row_data = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=rows[row_idx + i], column=col).value
                    # C列のフォーマットを調整（日付型の場合）
                    if col == 3 and isinstance(cell_value, datetime):
                        cell_value = cell_value.strftime("%Y/%m/%d")
                    row_data.append(cell_value)
                new_ws.append(row_data)

        # 次の4行を確認
        row_idx += 1

# 色を付けたファイルを保存する
output_path = "files/colored_sequence_PLOT.xlsx"
wb.save(output_path)

# シーケンスに一致した行だけを抜き出してファイル保存する
new_output_path = "files/colored_sequence_PLOT2.xlsx"
new_wb.save(new_output_path)
