# step1 PLOT用.xlsxを使って正常サイクルに色を付ける
#     色を付けたファイルをcolored_sequence_PLOT.xlsx保存する
#     正常サイクルだけを抜き出したファイルをcolored_sequence_PLOT2保存する
#     サイクルタイム
# step2 colored_sequence_PLOT2.xlsxを使って開始時間終了時間CTタイムの一覧表を作成する
#      colored_sequence_PLOT3
# step3 colored_sequence_PLOT3を使って画面から日付検索して表示する

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl
from datetime import datetime, timedelta
from openpyxl import Workbook
import tkinter as tk
from tkinter import Label, Entry, Button
from PIL import Image, ImageTk
import requests
from io import BytesIO


def step1():
    wb = load_workbook("files/PLOT用.xlsx")
    ws = wb.active

    # 新しいワークブックを作成
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # シーケンスの色を定義
    sequence_color = "FFFF00"  # 黄色

    # 必要なシーケンスを定義
    required_sequence = ["巻_1", "巻_2", "切_1", "切_2-1"]

    # C列の日付ごとに処理するための辞書を作成
    date_groups = {}
    for row in range(2, ws.max_row + 1):
        date_value = ws.cell(row=row, column=3).value  # C列（3列目）の日付を取得
        if isinstance(date_value, datetime):  # 日付型かどうかを確認
            date_value = date_value.date()  # 日付部分のみを取得（時間情報は無視）
            if date_value not in date_groups:
                date_groups[date_value] = []  # 新しい日付があればキーとして登録
            date_groups[date_value].append(row)  # 日付に対応する行を追加

    # 各日付ごとに処理を実行
    for date, rows in date_groups.items():
        row_idx = 0
        while row_idx <= len(rows) - 4:  # 最低3行あるか確認
            # 4行のシーケンスを取得
            current_sequence = [ws.cell(row=rows[row_idx + i], column=1).value for i in range(4)]

            # シーケンスが指定されたものと一致するかチェック
            if current_sequence == required_sequence:
                # 一致した場合、その4行に色を適用
                fill = PatternFill(start_color=sequence_color, end_color=sequence_color, fill_type="solid")
                for i in range(4):
                    for cell in ws[rows[row_idx + i]]:
                        cell.fill = fill

                # 新しいワークブックに一致した行をコピー
                for i in range(4):
                    row_data = []
                    for col in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=rows[row_idx + i], column=col).value
                        # C列のフォーマットを調整（日付型の場合）
                        if col == 3 and isinstance(cell_value, datetime):
                            cell_value = cell_value.strftime("%Y/%m/%d")
                        row_data.append(cell_value)
                    new_ws.append(row_data)

            # 次の4行を確認
            row_idx += 1

    # 色を付けたファイルを保存する
    output_path = "files/colored_sequence_PLOT.xlsx"
    wb.save(output_path)

    # シーケンスに一致した行だけを抜き出してファイル保存する
    new_output_path = "files/colored_sequence_PLOT2.xlsx"
    new_wb.save(new_output_path)


def step2():
    # 元データのExcelファイルを読み込む
    source_wb = load_workbook("files/colored_sequence_PLOT2.xlsx")
    source_ws = source_wb.active
    # 新しいExcelファイルを作成
    new_wb = Workbook()
    new_ws = new_wb.active
    # 転記先の表のヘッダーを設定
    new_ws.append(["ID", "スタート時間", "終了時間", "サイクルタイム", "日付"])
    # データを転記する巻１の時間を開始時間と切2‐1の時間を終了時間にする
    row_id = 1
    for i in range(1, source_ws.max_row + 1, 4):
        start_time = source_ws.cell(row=i, column=2).value
        end_time = source_ws.cell(row=i + 3, column=2).value
        cycle_date = source_ws.cell(row=i, column=3).value

        # Convert the start_time and end_time to seconds since midnight
        # 時間を秒に変換する
        start_seconds = start_time.hour * 3600 + start_time.minute * 60 + start_time.second
        end_seconds = end_time.hour * 3600 + end_time.minute * 60 + end_time.second

        # サイクルタイムの計算をする秒[終了時間-開始時間]
        time_difference_seconds = end_seconds - start_seconds

        # 差がマイナス表示の時は
        if time_difference_seconds < 0:
            time_difference_seconds += 86400  # 24時間86400秒をプラスする

        # Convert to the desired "MM:SS" format 割り切れたら分表示。余りは秒
        minutes, seconds = divmod(time_difference_seconds, 60)
        cycle_time = f"{minutes}:{seconds:02}"

        # Add to the new worksheet
        new_ws.append([row_id, start_time, end_time, cycle_time, cycle_date])
        row_id += 1
    # 新しいファイルに保存
    new_output_path = "files/colored_sequence_PLOT3.xlsx"
    new_wb.save(new_output_path)


def step3():

    def show_cycle_info():
        name = name_entry.get()

        standard_time = int(standard_entry.get())
        # エクセルを読み込む
        wb = openpyxl.load_workbook("files/colored_sequence_PLOT3.xlsx")
        ws = wb.active
        # ユーザーから日付を入力してもらう
        input_date = name
        # 現在の時間を取得する 午前０時からの経過時間 分表示
        now = datetime.now()
        print(now)
        midnight = datetime.combine(now.date(), datetime.min.time())
        # 経過時間を計算する
        elapsed_minutes = now - midnight

        # 停止時間は画面から入力
        stop_minutes = int(time_entry.get())
        # 稼働時間を計算
        # input_timeをtimedelta（分）として扱い、経過時間から引く
        adjusted_time = elapsed_minutes - timedelta(minutes=stop_minutes)
        print(adjusted_time)
        # 経過時間を秒に換算する
        elapsed_seconds = adjusted_time.total_seconds()
        # 稼働時間を standard_timeで割る
        result = elapsed_seconds / standard_time

        # #1 入力された日付のデータを抽出
        date_filter = []

        # 行数を取得する
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[4] == input_date:
                date_filter.append(row)
                count += 1

        if date_filter:
            # サイクルタイムを秒に変換してリストに格納
            cycle_times = [time_to_seconds(row[3]) for row in date_filter]
            # #3 ４列目の最大値を求める
            max_cycle_time = max(cycle_times)
            # #4 ４列目の最小値を求める
            min_cycle_time = min(cycle_times)
            # #5 ４列目の平均値を求める
            average_cycle_time = sum(cycle_times) / len(cycle_times)

            # サイクル数、最大、最小、平均を表示
            count_label.config(text=f"サイクル数: {count}")
            max_label.config(text=f"最大サイクルタイム:{max_cycle_time // 60}分{max_cycle_time % 60}秒")
            min_label.config(text=f"最小サイクルタイム: {min_cycle_time // 60}分{min_cycle_time % 60}秒")
            ave_label.config(
                text=f"平均サイクルタイム:{average_cycle_time // 60:.0f}分{average_cycle_time % 60:.0f}秒"
            )
            goal_label.config(text=f"目標サイクル数は：{int(result)}回")
        else:
            count_label.config(text=f"日付 '{count}' は見つかりませんでした。")
            max_label.config(text="")
            min_label.config(text="")
            ave_label.config(text="")

    # 秒を〇分〇秒にするための計算
    def time_to_seconds(time_str):
        minutes, seconds = map(int, time_str.split(":"))
        return minutes * 60 + seconds

    # GUIの初期設定

    root = tk.Tk()
    root.title("サイクル情報表示")

    # 現在の日付を取得
    today_date = datetime.now().strftime("%Y/%m/%d")

    # 日付入力フィールド
    Label(root, text="検索したい日付を入力してください:", font=("Arial", 16)).pack(padx=20, pady=10)
    name_entry = Entry(root, font=("Arial", 14))
    # insert で画面を開いたときに今日の日付を代入する
    name_entry.insert(0, today_date)
    name_entry.pack(padx=20, pady=10)

    Label(root, text="停止時間を入力してください（分）:", font=("Arial", 16)).pack(padx=20, pady=10)
    time_entry = Entry(root, font=("Arial", 14))
    time_entry.pack(padx=20, pady=10)

    Label(root, text="基準サイクルタイムを入力してください（秒）:", font=("Arial", 16)).pack(padx=20, pady=10)
    standard_entry = Entry(root, font=("Arial", 14))
    standard_entry.pack(padx=20, pady=10)

    # 表示ボタン

    Button(root, text="表示", font=("Arial", 14), command=show_cycle_info).pack(padx=20, pady=10)

    # サイクルタイム表示用ラベル（文字なし（中身なし））

    count_label = Label(root, font=("Arial", 20))
    count_label.pack(padx=20, pady=10)
    max_label = Label(root, font=("Arial", 16))
    max_label.pack(padx=20, pady=5)
    min_label = Label(root, font=("Arial", 16))
    min_label.pack(padx=20, pady=5)
    ave_label = Label(root, font=("Arial", 16))
    ave_label.pack(padx=20, pady=5)
    goal_label = Label(root, font=("Arial", 20))
    goal_label.pack(padx=20, pady=10)

    # GUIの表示
    root.mainloop()


#
if __name__ == "__main__":
    step1()
    step2()
    step3()
