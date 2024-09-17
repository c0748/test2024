import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime, timedelta

# 元データのExcelファイルを読み込む
source_wb = load_workbook("files/colored_sequence_PLOT2.xlsx")
source_ws = source_wb.active
# 新しいExcelファイルを作成
new_wb = Workbook()
new_ws = new_wb.active
# 転記先の表のヘッダーを設定
new_ws.append(["ID", "スタート時間", "終了時間", "サイクルタイム"])
# データを転記
row_id = 1
for i in range(1, source_ws.max_row + 1, 4):
    start_time = source_ws.cell(row=i, column=2).value
    end_time = source_ws.cell(row=i + 3, column=2).value

    # Convert the start_time and end_time to seconds since midnight
    # 時間を秒に変換する
    start_seconds = start_time.hour * 3600 + start_time.minute * 60 + start_time.second
    end_seconds = end_time.hour * 3600 + end_time.minute * 60 + end_time.second

    # Calculate the difference in seconds
    time_difference_seconds = end_seconds - start_seconds

    # 差がマイナス表示の時は
    if time_difference_seconds < 0:
        time_difference_seconds += 86400  # 24時間86400秒をプラスする

    # Convert to the desired "MM:SS" format 割り切れたら分表示。余りは秒
    minutes, seconds = divmod(time_difference_seconds, 60)
    cycle_time = f"{minutes}:{seconds:02}"

    # Add to the new worksheet
    new_ws.append([row_id, start_time, end_time, cycle_time])
    row_id += 1
# 新しいファイルに保存
new_output_path = "files/colored_sequence_PLOT3.xlsx"
new_wb.save(new_output_path)
