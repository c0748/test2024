import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the workbook and select the active worksheet
wb = load_workbook("files/colored_sequence_PLOT.xlsx")
ws = wb.active

new_wb = openpyxl.Workbook()
new_ws = new_wb.active
sequence_color = "FFFF00"  # Yellow

for row in ws.iter_rows():
    # A列の色をチェック
    if row[0].fill.start_color.index == sequence_color:
        new_ws.append([cell.value for cell in row])

# 新しいワークブックを保存
output_path = "files/colored_sequence_PLOT2.xlsx"
new_wb.save(output_path)
