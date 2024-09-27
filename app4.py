#このファイルは使わない
# app10.pyに改訂済み

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import openpyxl
from datetime import datetime

# ワークブックを読み込みアクティブなワークシートを選択
wb = load_workbook("files/PLOT用.xlsx")
ws = wb.active

# 新しいワークブックを作成
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# シーケンス用の色を定義
sequence_color = "FFFF00"  # Yellow
# 必要なシーケンスを定義
required_sequence = ["巻_1", "巻_2", "切_1", "切_2-1"]
# ４行単位でブロックを確認するために行をループ処理
row = 2
while row <= ws.max_row - 3:  # 残り３行になるまで処理を繰り返す
    current_sequence = [ws.cell(row=row + i, column=1).value for i in range(4)]
    # シーケンス内のセル画からの場合は処理を終了する
    if any(value is None for value in current_sequence):
        break
    # 現在のシーケンスが定義されたシーケンスと一致するか確認
    if current_sequence == required_sequence:
        # 一致した行に色を適用
        # このループは、現在の行 (row) から始まる4行を処理します。i は 0 から 3 までの値を取ります。
        fill = PatternFill(start_color=sequence_color, end_color=sequence_color, fill_type="solid")
        for i in range(4):
            for cell in ws[row + i]:
                cell.fill = fill
        for i in range(4):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row + i, column=col).value
                if col == 3 and isinstance(cell_value, datetime):
                    cell_value = cell_value.strftime("%Y/%m/%d")
                row_data.append(cell_value)
            new_ws.append(row_data)
    # 次の行に移動する
    row += 1
# 修正したExcelファイルを保存する
output_path = "files/colored_sequence_PLOT.xlsx"
wb.save(output_path)
# 色が適用された元のエクセルファイルを保存する
new_output_path = "files/colored_sequence_PLOT2.xlsx"
new_wb.save(new_output_path)
