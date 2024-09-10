from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the workbook and select the active worksheet
wb = load_workbook("files/PLOT用.xlsx")
ws = wb.active
# Define color for the sequence
sequence_color = "FFFF00"  # Yellow
# Define the required sequence
required_sequence = ["巻_1", "巻_2", "切_1", "切_2-1"]
# Loop through the rows to check 4-row blocks
row = 2
while row <= ws.max_row - 3:  # Ensure we have at least 4 rows left to check
    current_sequence = [ws.cell(row=row + i, column=1).value for i in range(4)]
    # If any cell in the sequence is empty, break the loop
    if any(value is None for value in current_sequence):
        break
    # Check if the current sequence matches the required sequence
    if current_sequence == required_sequence:
        # Apply color to the matching rows
        fill = PatternFill(start_color=sequence_color, end_color=sequence_color, fill_type="solid")
        for i in range(4):
            for cell in ws[row + i]:
                cell.fill = fill
    # Move to the next row to check the next 4-row block
    row += 1
# Save the modified Excel file
output_path = "files/colored_sequence_PLOT.xlsx"
wb.save(output_path)
